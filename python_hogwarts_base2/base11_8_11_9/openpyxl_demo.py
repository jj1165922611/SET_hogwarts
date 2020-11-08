#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time       : 2020-11-05
# @Author     : Joey Jiang
# @File       : openpyxl_demo.py
# @Software   : PyCharm
# @Description: openpyxl读写excel

from openpyxl import Workbook

# 1、写入并创建excel
wb=Workbook()
ws=wb.active
ws['A1']="身高"
ws['B1']="体重"
ws['C1']="健康体重"
ws['D1']="是否健康"
height=[170,165,185]
weight=[60,50,70]
for i in range(len(height)):
    ws.cell(i+2,1,height[i])
    ws.cell(i+2,2,weight[i])
    health_weight=(height[i]-70)*0.6
    ws.cell(i+2,3,int(health_weight))
    if weight[i]<=int(health_weight):
        ws.cell(i+2,4,"健康")
wb.save("sample.xlsx")
