#!usr/bin/env python
# -*- coding: utf-8 -*-
# @Time       : 2020-05-07
# @Author     : Joey Jiang
# @File       : excel_openpyxl.py
# @Software   : Visual Studio Code
# @Description: 使用openpyxl实现excel写入

from openpyxl import load_workbook

'''
定义一个类，类名为PracticeExcel
'''
class PracticeExcel:

    '''
    *初始化类PracticeExcel
    *参数：
        data: 字典类型，key是身高，value是体重
    '''
    def __init__(self,data:dict):
        self.data=data # 定义一个变量data，将参数data的值传给它
    
    '''
    *创建一个方法write_to_excel
    *参数:
        file_name: 文件名
        sheet_name: sheet名
    '''
    def write_to_excel(self,file_name,sheet_name):
        data=self.data # 定义一个变量data，将self.data的值赋给它
        # 1.加载指定文件
        wb=load_workbook(file_name) # 定义一个变量wb，将指定文件的内容赋值给它
        # 2.加载指定sheet
        sheet=wb[sheet_name] # 定义变量sheet，将指定sheet的内容传给它
        # 3.写入标题
        sheet['A1']="身高" # 在sheet的A1位置写入文字身高
        sheet['B1']="体重" # 在sheet的B1位置写入文字体重
        sheet['C1']="备注" # 在sheet的C1位置写入文字备注
        # 4.写入数据
        print(type(data.keys())) # 打印出字典的keys方法返回值的类型
        height_list=list(data.keys()) # 定义一个变量height_list，将data中的关键字转换成list类型后赋值给它
        # 4.1身高和体重的数据直接写入
        for i in range(len(height_list)): # 定义一个循环，循环的次数为height_list列表的长度，变量为i
            sheet.cell(row=i+2,column=1).value=height_list[i] # 将height_list列表中索引为i的元素写入到A列i+2行
            sheet.cell(row=i+2,column=2).value=data.get(height_list[i])# 将data中关键字为height_list[i]的值写入到B列i+2行
        # 4.2只有健康体重需要在旁边备注:健康体重计算公式：（身高cm-70）×60%
            if (int(height_list[i])-70)*0.6 ==int(data.get(height_list[i]))*1.0: # 判断体重是否为健康体重
                sheet.cell(row=i+2,column=3).value="健康体重" # 将文字“健康体重”写入到C列i+2行
        # 5.保存指定文件
        wb.save(file_name) # 保存指定文件

if __name__ == "__main__":  
    file_name="./sample.xlsx" # 定义一个变量file_name，将指定文件名称赋值给它
    sheet_name="Sheet" # 定义一个变量sheet_name，将指定sheet名称赋值给它
    data={"180":"57","190":"72","200":"95"} # 定义一个字典data，关键字为身高，值为体重
    PracticeExcel(data).write_to_excel(file_name,sheet_name) # 使用类对象，传入参数data，调用write_to_excel方法