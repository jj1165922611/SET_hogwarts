#!usr/bin/env python
# -*- coding: utf-8 -*-
# @Time       : 2020-05-07
# @Author     : Joey Jiang
# @File       : base8_7_2.py
# @Software   : Visual Studio Code
# @Description: python外部数据源文件处理

from openpyxl import Workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()

# Save the file
wb.save("sample.xlsx")